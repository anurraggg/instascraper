"""
Instagram Scraper - CUSTOM FINAL VERSION
Based on scraper_ready.py logic
"""

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook, Workbook
# import pandas as pd # Disabled due to env issues
import time
import re
import random
import os
import json
from pathlib import Path
from datetime import datetime
import pickle

def save_cookies(driver, path):
    try:
        with open(path, 'wb') as filehandler:
            pickle.dump(driver.get_cookies(), filehandler)
        print(f"    🍪 Cookies saved to {path}")
    except Exception as e:
        print(f"    ⚠️ Failed to save cookies: {e}")

def load_cookies(driver, path):
    try:
        if os.path.exists(path):
            with open(path, 'rb') as cookiesfile:
                cookies = pickle.load(cookiesfile)
                for cookie in cookies:
                    driver.add_cookie(cookie)
            print(f"    🍪 Cookies loaded from {path}")
            return True
    except Exception as e:
        print(f"    ⚠️ Failed to load cookies: {e}")
    return False

def extract_number(text):
    """Extract number from text."""
    if not text:
        return 0
    text = str(text).strip().upper().replace(',', '')
    multiplier = 1
    if 'K' in text:
        multiplier = 1000
        text = text.replace('K', '')
    elif 'M' in text:
        multiplier = 1000000
        text = text.replace('M', '')
    elif 'B' in text:
        multiplier = 1000000000
        text = text.replace('B', '')
    match = re.search(r'[\d.]+', text)
    if match:
        try:
            return int(float(match.group()) * multiplier)
        except:
            return 0
    return 0

def is_valid_username(username):
    if not username:
        return False
    ignored = ['home', 'explore', 'reels', 'direct', 'accounts', 'create', 'profile', 'stories', 'instagram', 'search', 'notifications', 'more', 'settings', 'messages', 'p', 'reel']
    return username.lower() not in ignored and len(username) > 1

def get_follower_count(driver, username):
    """Robustly extract follower count from a user profile."""
    try:
        driver.execute_script(f"window.open('https://www.instagram.com/{username}/', '_blank');")
        time.sleep(2)
        driver.switch_to.window(driver.window_handles[-1])
        time.sleep(5)
        
        followers = 0
        
        # Strategy 1: Meta Tag (Fastest & Most Reliable)
        try:
            meta_desc = driver.find_element(By.XPATH, "//meta[@property='og:description']").get_attribute("content")
            if meta_desc:
                # Format: "1M Followers, 200 Following, 50 Posts..."
                followers_text = meta_desc.split('Followers')[0].strip()
                followers = extract_number(followers_text)
                if followers > 0:
                    driver.close()
                    driver.switch_to.window(driver.window_handles[0])
                    return followers
        except: pass

        # Strategy 2: Title Tag
        try:
            title_text = driver.title
            # Format: "Name (@username) • Instagram photos and videos" - sometimes contains stats
            if 'Followers' in title_text:
                f_text = title_text.split('Followers')[0].strip().split()[-1]
                followers = extract_number(f_text)
                if followers > 0:
                    driver.close()
                    driver.switch_to.window(driver.window_handles[0])
                    return followers
        except: pass

        # Strategy 3: Page Selectors (Visible Elements)
        print("    ↳ Meta/Title failed, checking page elements...")
        follower_selectors = [
            "//a[contains(@href, '/followers/')]/span",
            "//ul//li[2]//span//span",
            "//span[contains(text(), 'followers')]",
            "//div[contains(text(), 'followers')]"
        ]
        
        for sel in follower_selectors:
            try:
                elem = driver.find_element(By.XPATH, sel)
                val = extract_number(elem.get_attribute("title") or elem.text)
                if val > 0:
                    followers = val
                    break
            except: continue
        
        driver.close()
        driver.switch_to.window(driver.window_handles[0])
        return followers
        
    except Exception as e:
        print(f"    ⚠️ Follower func error: {e}")
        try:
            if len(driver.window_handles) > 1:
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
        except: pass
        return 0

def handle_popups(driver):
    """Check for and dismiss common Instagram popups."""
    try:
        # 1. "Never miss a post" / Login popup (Close button)
        close_btns = driver.find_elements(By.XPATH, "//div[@role='dialog']//*[name()='svg' and @aria-label='Close']")
        for btn in close_btns:
            print("    found popup close button, clicking...")
            driver.execute_script("arguments[0].click();", btn)
            time.sleep(1)
            return True
            
        # 2. "Not Now" buttons
        not_now = driver.find_elements(By.XPATH, "//button[contains(text(), 'Not Now') or contains(text(), 'Not now')]")
        if not_now:
            print("    found 'Not Now', clicking...")
            not_now[0].click()
            time.sleep(1)
            return True
            
        return False
    except: return False

def main(input_file='input/INSTAURL.xlsx', output_file='output/instagram_fresh_results.xlsx'):
    print("\n" + "="*60)
    print("INSTAGRAM CUSTOM FINAL SCRAPER")
    print("="*60)

    # Load URLs
    INPUT_FILE = input_file
    try:
        wb = load_workbook(INPUT_FILE)
        ws = wb.active
        urls = []
        # Assuming header in row 1, data starts row 2
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                val = str(row[0]).strip()
                if val.startswith('http'):
                    urls.append(val)
        print(f"\n✓ Loaded {len(urls)} URLs from {INPUT_FILE}")
    except Exception as e:
        print(f"✗ Error reading input: {e}")
        return

    # Setup Chrome
    print("\n🚀 Starting Chrome...")
    options = Options()
    options.add_argument('--log-level=3')
    options.add_experimental_option("excludeSwitches", ["enable-logging"])
    # options.add_argument('--headless') # Keep it visible to monitor

    driver = webdriver.Chrome(options=options)
    driver.maximize_window()
    print("✅ Chrome ready")

    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC

    # Login Logic
    print("\n🔐 Checking session...")
    driver.get("https://www.instagram.com/")
    time.sleep(3)
    
    # Try to load cookies
    if load_cookies(driver, "cookies.pkl"):
        driver.refresh()
        time.sleep(5)
    
    is_logged_in = False
    wait = WebDriverWait(driver, 15)
    
    # Check if already logged in (restored session)
    try:
        wait.until(EC.presence_of_element_located((By.XPATH, "//svg[@aria-label='Home' or @aria-label='Home Feed']")))
        print("✓ Session restored successfully.")
        is_logged_in = True
    except:
        print("ℹ️ Session expired or not found. Performing full login...")
        driver.get("https://www.instagram.com/accounts/login/")

    if not is_logged_in:
        try:
            wait = WebDriverWait(driver, 15)
            user_input = wait.until(EC.presence_of_element_located((By.NAME, "username")))
            pass_input = driver.find_element(By.NAME, "password")
            
            user_input.send_keys(os.environ.get("IG_USERNAME", ""))
            pass_input.send_keys(os.environ.get("IG_PASSWORD", ""))
            pass_input.send_keys(Keys.ENTER)
            print("✓ Credentials entered. Waiting for login to complete...")
            time.sleep(10)
            
            # Check success
            try:
                wait.until(EC.presence_of_element_located((By.XPATH, "//svg[@aria-label='Home' or @aria-label='Home Feed']")))
                print("✓ Login successful.")
                is_logged_in = True
            except:
                # Handle "Save Info" popup
                try:
                    save_info = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[text()='Not now' or text()='Not Now']")))
                    save_info.click()
                    time.sleep(3)
                except: pass
                
                # Handle "Notifications" popup
                try:
                    notif = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[text()='Not Now']")))
                    notif.click()
                    time.sleep(3)
                except: pass
                
            # Save cookies if login worked (or we are in)
            if is_logged_in or len(driver.get_cookies()) > 0:
                 save_cookies(driver, "cookies.pkl")

        except Exception as e:
            print(f"⚠️ Automated login failed: {e}")
            
    # Final check
    try:
        wait.until(EC.presence_of_element_located((By.XPATH, "//svg[contains(@aria-label, 'Home')]")))
        print("✅ Login verified.")
        save_cookies(driver, "cookies.pkl")
    except:
        print("👉 Please ensure you are logged in in the browser window.")
        print("👉 Press ENTER to continue...")
        input()
        save_cookies(driver, "cookies.pkl")

    # Results tracking (Starting Fresh)
    results = []
    Path('output').mkdir(exist_ok=True)
    OUTPUT_FILE = output_file
    
    # User requested to start AFRESH, so we ignore existing results
    print("✨ Starting FRESH scraping...")

    for i in range(len(urls)):
        url = urls[i]
        print(f"\n[{i+1}/{len(urls)}] {url[:60]}...")
        
        if "youtube.com" in url or "youtu.be" in url:
            print("  ⚠️ Skipping YouTube URL.")
            results.append({'URL': url, 'Status': 'Skipped (YouTube)'})
            continue

        try:
            driver.get(url)
            time.sleep(8)
            
            # Handle any popping up walls
            handle_popups(driver)
            
            # Check for Login Wall explicitly
            if "Login • Instagram" in driver.title:
                print(f"  ❌ Login Wall detected! Pausing for manual login...")
                print("👉 Please Log In manually in the browser window.")
                print("👉 Press ENTER here when you are done...")
                input()
                save_cookies(driver, "cookies.pkl")
                # Retry loading the page
                driver.get(url) 
                time.sleep(8)
                handle_popups(driver)

            # Debug: Check page content
            print(f"  🔍 Page Title: {driver.title}")
            
            # Check for "Post unavailable"
            if "isn't available" in driver.title or "Post unavailable" in driver.page_source or "Login • Instagram" in driver.title:
                print(f"  ❌ Post unavailable or Login wall! Title: {driver.title}")
                result = {'URL': url, 'Status': 'Unavailable/Login Required'}
                results.append(result)
                continue

            # Metrics initialization
            likes = 0
            comments = 0
            hashtags = []
            followers = 0
            username = ''

            # 1. Extraction from Post Page
            try:
                # Likes
                selectors_likes = [
                    "//a[contains(@href, '/liked_by/')]//span",
                    "//section//span[contains(text(), 'likes')]",
                    "//div[contains(text(), 'others')]",
                    "//span[contains(@aria-label, 'like')]//following-sibling::span"
                ]
                for sel in selectors_likes:
                    try:
                        elem = driver.find_element(By.XPATH, sel)
                        val = extract_number(elem.text)
                        if val > 0: likes = val; break
                    except: continue
                
                # Comments
                selectors_comments = [
                    "//span[contains(text(), 'comments')]",
                    "//div[contains(@aria-label, 'Comment')]//following-sibling::span",
                    "//ul//li[contains(text(), 'comments')]",
                    # NEW: Generic number next to Comment icon (for both Reel and Post view)
                    "//*[name()='svg' and @aria-label='Comment']/ancestor::div[@role='button']//span",
                    "//*[name()='svg' and @aria-label='Comment']/../following-sibling::span",
                    "//*[name()='svg' and @aria-label='Comment']/ancestor::span/following-sibling::span", 
                    "//section//*[name()='svg' and @aria-label='Comment']/ancestor::div[2]//span", 
                ]
                for sel in selectors_comments:
                    try:
                        elem = driver.find_element(By.XPATH, sel)
                        val = extract_number(elem.text)
                        if val > 0: comments = val; break
                    except: continue

                # Likes - Enhanced for Reels
                selectors_likes = [
                    "//a[contains(@href, '/liked_by/')]//span",
                    "//section//span[contains(text(), 'likes')]",
                    "//div[contains(text(), 'others')]",
                    "//span[contains(@aria-label, 'like')]//following-sibling::span",
                     # NEW: Reel sidebar like count
                    "//button//div//*[name()='svg' and @aria-label='Like']/ancestor::button/following-sibling::span",
                    "//section//div//*[name()='svg' and @aria-label='Like']/ancestor::div//span",
                    # Generic icon neighbor
                    "//*[name()='svg' and @aria-label='Like']/../following-sibling::span",
                    "//*[name()='svg' and @aria-label='Like']/ancestor::div[@role='button']//span",
                ]
                for sel in selectors_likes:
                    try:
                        elem = driver.find_element(By.XPATH, sel)
                        val = extract_number(elem.text)
                        if val > 0: likes = val; break
                    except: continue
                
                # Hashtags & Caption
                try:
                    caption_elem = driver.find_element(By.XPATH, "//h1 | //title/following-sibling::*//span[contains(text(), '#')]")
                    caption_text = caption_elem.text
                    hashtags = re.findall(r'#\w+', caption_text)
                except:
                    # Fallback caption check
                    try:
                        body_text = driver.find_element(By.TAG_NAME, "body").text
                        hashtags = re.findall(r'#\w+', body_text[:2000])
                    except: pass
                
                # Username
                try:
                    user_elem = driver.find_element(By.XPATH, "//header//a[contains(@class, 'x1i10hfl')]")
                    username = user_elem.get_attribute("href").strip('/').split('/')[-1]
                except:
                    title = driver.title
                    if '•' in title: username = title.split('•')[0].strip()
                    elif '(' in title: username = title.split('(')[0].strip()

                # 2. Visit Creator Profile for Follower Count
                # REEL SPECIFIC LOGIC: If username not found, try to find it in Reel UI
                if not username and "/reel/" in url:
                    try:
                        # Try finding username in Reel layout
                        user_elem = driver.find_element(By.XPATH, "//div[contains(@class, '_aaq8')]//a[contains(@href, '/')]") 
                        username = user_elem.get_attribute("href").strip('/').split('/')[-1]
                    except: pass
                
                # REEL SPECIFIC LOGIC: Fallback to /p/ if standard extraction failed
                if (likes == 0 or comments == 0) and "/reel/" in url:
                    print("    Checking /p/ fallback for Reel...")
                    p_url = url.replace("/reel/", "/p/")
                    driver.get(p_url)
                    time.sleep(5)
                    handle_popups(driver)
                    
                    # Retry extraction on /p/ page
                    # Likes
                    for sel in selectors_likes:
                         try:
                             elem = driver.find_element(By.XPATH, sel)
                             val = extract_number(elem.text)
                             if val > 0: likes = val; break
                         except: continue
                    
                    # Comments
                    for sel in selectors_comments:
                        try:
                            elem = driver.find_element(By.XPATH, sel)
                            val = extract_number(elem.text)
                            if val > 0: comments = val; break
                        except: continue

                    # Username fallback
                    if not username:
                        try:
                            user_elem = driver.find_element(By.XPATH, "//header//a[contains(@class, 'x1i10hfl')]")
                            username = user_elem.get_attribute("href").strip('/').split('/')[-1]
                        except: pass

                if is_valid_username(username):
                    print(f"  🕵️ Visiting profile for followers: {username}")
                    followers = get_follower_count(driver, username)
                    print(f"  ✓ Followers: {followers:,}")
            
            except Exception as e:
                print(f"  ⚠️ Extraction error: {e}")
                try:
                    driver.save_screenshot(f"output/debug_error_{i}.png")
                except: pass

            # DEBUG: Screenshot if failed (or partial fail)
            if likes == 0 or comments == 0:
                print("  ⚠️ Missing metrics (Likes or Comments). Saving debug screenshot...")
                try:
                    screenshot_path = f"output/debug_fail_{i}.png"
                    driver.save_screenshot(screenshot_path)
                    print(f"  📸 Saved to {screenshot_path}")
                except: pass

            result = {
                'URL': url,
                'Username': username,
                'Followers': followers,
                'Likes': likes,
                'Comments': comments,
                'Hashtags': ", ".join(hashtags) if hashtags else "",
                'Status': 'Success',
                'Timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            results.append(result)
            print(f"  📊 Success! Followers: {followers:,} | Likes: {likes:,} | Comments: {comments:,} | Hashtags: {len(hashtags)}")
            # EXTRA DEBUG PRINT
            print(f"  [DEBUG] Captured: Likes={likes}, Comments={comments}, User={username}")

        except Exception as e:
            print(f"  ✗ Error: {e}")
            results.append({'URL': url, 'Status': 'Error', 'Error': str(e)})

        # Periodic save (every row for safety in fresh start)
        # Periodic save (every row for safety in fresh start)
        try:
            wb_out = Workbook()
            ws_out = wb_out.active
            # Headers
            headers = ['URL', 'Username', 'Followers', 'Likes', 'Comments', 'Hashtags', 'Status', 'Timestamp', 'Error']
            ws_out.append(headers)
            # Data
            for res in results:
                row = [
                    res.get('URL', ''),
                    res.get('Username', ''),
                    res.get('Followers', 0),
                    res.get('Likes', 0),
                    res.get('Comments', 0),
                    res.get('Hashtags', ''),
                    res.get('Status', ''),
                    res.get('Timestamp', ''),
                    res.get('Error', '')
                ]
                ws_out.append(row)
            wb_out.save(OUTPUT_FILE)
        except Exception as e:
            print(f"  ⚠️ Error saving output: {e}")

    driver.quit()
    print("\n✅ ALL DONE! Fresh results in: " + OUTPUT_FILE)

if __name__ == "__main__":
    main()
